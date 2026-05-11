# ================================================================
# IIAP UploadFiles - Production-Grade Locust Load Test
# locustfile.py lives at: UploadFiles/ (project root)
#
# REALISTIC SCENARIO:
#   350 candidates each upload 3 files (Word, Excel, PPT)
#   149 staff browse labs and print files
#   1 admin does oversight
#
# QUICK START:
#   pip install locust openpyxl
#   cd UploadFiles/backend && python loadtest/seed_test_data.py
#   cd UploadFiles/backend && python manage.py runserver
#   cd UploadFiles && locust -f locustfile.py --host http://localhost:8000
#   Open http://localhost:8089
#
# RECOMMENDED SETTINGS IN LOCUST UI:
#   Users     : 356  (350 candidates + 5 staff + 1 admin)
#   Spawn rate: 10   (ramp up over ~35 seconds - realistic exam start)
#
# CREDENTIALS: read from backend/loadtest/test_users.xlsx
#   Columns: SI No | UserID | Name | Role | Password | Mobile | Lab | Action
# ================================================================

from __future__ import annotations

import os
import random
import threading
from collections import defaultdict
from typing import Optional

from locust import HttpUser, between, events, tag, task
from locust.exception import StopUser

try:
    from openpyxl import load_workbook
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# ── Paths ──────────────────────────────────────────────────────
ROOT       = os.path.dirname(os.path.abspath(__file__))
SAMPLE_DIR = os.path.join(ROOT, "backend", "loadtest", "sample_files")
XLSX_PATH  = os.path.join(ROOT, "backend", "loadtest", "test_users.xlsx")

SAMPLE_WORD  = os.path.join(SAMPLE_DIR, "test.docx")
SAMPLE_EXCEL = os.path.join(SAMPLE_DIR, "test.xlsx")
SAMPLE_PPT   = os.path.join(SAMPLE_DIR, "test.pptx")


# ── Load credentials from Excel ────────────────────────────────
# Correct column order: SI No(0) | UserID(1) | Name(2) | Role(3)
#                       Password(4) | Mobile(5) | Lab(6) | Action(7)
def _load_creds(xlsx_path: str) -> dict:
    by_role: dict[str, list[dict]] = defaultdict(list)

    if _HAS_OPENPYXL and os.path.exists(xlsx_path):
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb["All Users"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[1] or row[1] == "TOTAL":
                continue
            # col 0=SI, 1=UserID, 2=Name, 3=Role, 4=Password, 5=Mobile, 6=Lab, 7=Action
            user_id  = str(row[1]).strip()
            role     = str(row[3]).strip()      # Candidate / Staff / Admin
            password = str(row[4]).strip()
            lab      = str(row[6]).strip() if row[6] else ""
            if role in ("Candidate", "Staff", "Admin"):
                by_role[role].append({
                    "user_id":  user_id,
                    "password": password,
                    "lab":      lab,
                })
        wb.close()
        total = sum(len(v) for v in by_role.values())
        print(
            f"[locust] Loaded {total} credentials from Excel: "
            f"{len(by_role['Candidate'])} candidates, "
            f"{len(by_role['Staff'])} staff, "
            f"{len(by_role['Admin'])} admins"
        )
    else:
        # Inline fallback — matches test_users.xlsx content
        if not _HAS_OPENPYXL:
            print("[locust] WARNING: openpyxl not installed — using inline credentials")
        else:
            print(f"[locust] WARNING: {xlsx_path} not found — using inline credentials")
        for i in range(1, 351):
            lab = "Lab-" + chr(65 + ((i - 1) // 50))
            by_role["Candidate"].append({"user_id": f"LTCAND{i:03d}", "password": "Test@1234", "lab": lab})
        for i in range(1, 150):
            lab = "Lab-" + chr(65 + ((i - 1) // 21))
            by_role["Staff"].append({"user_id": f"LTSTAFF{i:03d}", "password": "Test@1234", "lab": lab})
        by_role["Admin"].append({"user_id": "LTADMIN01", "password": "Test@1234", "lab": "All"})

    return dict(by_role)


_CREDS          = _load_creds(XLSX_PATH)
CANDIDATE_CREDS = _CREDS.get("Candidate", [])
STAFF_CREDS     = _CREDS.get("Staff", [])
ADMIN_CREDS     = _CREDS.get("Admin", [])

# Verify sample files exist
for _fp in (SAMPLE_WORD, SAMPLE_EXCEL, SAMPLE_PPT):
    if not os.path.exists(_fp):
        print(f"[locust] WARNING: sample file missing → {_fp}")
        print(f"[locust]          Create any small file at that path to fix.")


# ── File cache — read each sample once, share across threads ──
_file_cache: dict[str, bytes] = {}
_cache_lock = threading.Lock()

def _read_file(path: str) -> Optional[bytes]:
    if not os.path.exists(path):
        return None
    with _cache_lock:
        if path not in _file_cache:
            with open(path, "rb") as fh:
                _file_cache[path] = fh.read()
    return _file_cache[path]


# ── Login helper ───────────────────────────────────────────────
def do_login(client, cred_list: list) -> tuple:
    """Login with a random credential. Returns (access_token, user_id)."""
    cred = random.choice(cred_list)
    with client.post(
        "/api/auth/login/",
        json={"user_id": cred["user_id"], "password": cred["password"]},
        name="POST /api/auth/login/",
        catch_response=True,
    ) as resp:
        if resp.status_code == 200:
            data = resp.json()
            return data.get("access"), data.get("user_id")
        resp.failure(
            f"Login failed [{cred['user_id']}]: "
            f"{resp.status_code} — {resp.text[:80]}"
        )
        return None, None


# ── Upload helper ──────────────────────────────────────────────
def _upload(client, headers: dict, filepath: str, filename: str, file_type: str):
    """POST a file to /api/files/. Treats 201 and 403 as success."""
    data = _read_file(filepath)
    if data is None:
        return
    with client.post(
        "/api/files/",
        files={"file": (filename, data, "application/octet-stream")},
        data={"file_type": file_type},
        headers=headers,
        name=f"POST /api/files/ ({file_type})",
        catch_response=True,
    ) as resp:
        # 201 = new upload, 403 = already printed (lock), both expected
        if resp.status_code in (201, 403):
            resp.success()
        else:
            resp.failure(f"Upload {file_type} → {resp.status_code}: {resp.text[:60]}")


# ══════════════════════════════════════════════════════════════
# USER CLASS 1: CandidateUser
#
# weight=7 → 70% of all virtual users are Candidates
# wait_time between(5,20) → realistic exam pace
#
# REALISTIC FLOW:
#   Each candidate logs in once, then cycles through:
#   - Viewing their file list (most frequent — page auto-refreshes)
#   - Uploading Word, Excel, PPT files
#   - Checking session validity
# ══════════════════════════════════════════════════════════════
class CandidateUser(HttpUser):
    weight    = 7
    wait_time = between(5, 20)

    def on_start(self):
        """Login once. Stop this user if login fails."""
        token, uid = do_login(self.client, CANDIDATE_CREDS)
        if not token:
            raise StopUser()
        self.token   = token
        self.user_id = uid
        self.headers = {"Authorization": f"Bearer {token}"}

    def on_stop(self):
        """Blacklist token on logout."""
        self.client.post(
            "/api/auth/logout/",
            json={"refresh": ""},
            headers=self.headers,
            name="POST /api/auth/logout/ (candidate)",
        )

    @task(3)
    @tag("candidate", "read")
    def view_my_files(self):
        """
        GET own file list.
        weight=3 — most common action. The candidate page
        auto-refreshes every 2 minutes in the real app.
        """
        self.client.get(
            "/api/files/",
            headers=self.headers,
            name="GET /api/files/ (own list)",
        )

    @task(2)
    @tag("candidate", "upload")
    def upload_word(self):
        """Upload Word file. weight=2."""
        _upload(self.client, self.headers, SAMPLE_WORD, "exam_word.docx", "Word")

    @task(2)
    @tag("candidate", "upload")
    def upload_excel(self):
        """Upload Excel file. weight=2."""
        _upload(self.client, self.headers, SAMPLE_EXCEL, "exam_excel.xlsx", "Excel")

    @task(2)
    @tag("candidate", "upload")
    def upload_ppt(self):
        """Upload PPT file. weight=2."""
        _upload(self.client, self.headers, SAMPLE_PPT, "exam_ppt.pptx", "PPT")

    @task(1)
    @tag("candidate", "read")
    def check_session(self):
        """Verify JWT is still valid. weight=1."""
        self.client.get(
            "/api/auth/me/",
            headers=self.headers,
            name="GET /api/auth/me/ (candidate)",
        )


# ══════════════════════════════════════════════════════════════
# USER CLASS 2: StaffUser
#
# weight=2 → 20% of virtual users are Staff
# wait_time between(3,10) → staff works faster than candidates
#
# REALISTIC FLOW:
#   Staff logs in, selects a lab, browses candidates,
#   views files for printing. Repeat.
# ══════════════════════════════════════════════════════════════
class StaffUser(HttpUser):
    weight    = 2
    wait_time = between(3, 10)

    def on_start(self):
        token, _ = do_login(self.client, STAFF_CREDS)
        if not token:
            raise StopUser()
        self.token    = token
        self.headers  = {"Authorization": f"Bearer {token}"}
        self.cand_ids: list = []
        self.labs: list     = []
        self._load_labs()

    def on_stop(self):
        self.client.post(
            "/api/auth/logout/",
            json={"refresh": ""},
            headers=self.headers,
            name="POST /api/auth/logout/ (staff)",
        )

    def _load_labs(self):
        """Load lab list once at start."""
        try:
            resp = self.client.get(
                "/api/admin/labs/",
                headers=self.headers,
                name="GET /api/admin/labs/ (init)",
            )
            if resp.status_code == 200:
                self.labs = resp.json()
                if self.labs:
                    self._load_candidates(random.choice(self.labs))
        except Exception:
            pass

    def _load_candidates(self, lab: str):
        """Load candidates for a specific lab."""
        try:
            resp = self.client.get(
                f"/api/admin/labs/?lab={lab}",
                headers=self.headers,
                name="GET /api/admin/labs/?lab= (init)",
            )
            if resp.status_code == 200:
                self.cand_ids = [c["user_id"] for c in resp.json()]
        except Exception:
            pass

    @task(2)
    @tag("staff", "read")
    def list_labs(self):
        """GET all labs — staff opens the Print Files page."""
        with self.client.get(
            "/api/admin/labs/",
            headers=self.headers,
            name="GET /api/admin/labs/ (all)",
            catch_response=True,
        ) as resp:
            if resp.status_code == 200:
                self.labs = resp.json()

    @task(3)
    @tag("staff", "read")
    def list_candidates_in_lab(self):
        """GET candidates for a lab — staff selects a lab."""
        if not self.labs:
            return
        lab = random.choice(self.labs)
        with self.client.get(
            f"/api/admin/labs/?lab={lab}",
            headers=self.headers,
            name="GET /api/admin/labs/?lab= (candidates)",
            catch_response=True,
        ) as resp:
            if resp.status_code == 200:
                self.cand_ids = [c["user_id"] for c in resp.json()]

    @task(4)
    @tag("staff", "read")
    def view_candidate_files(self):
        """
        GET files for a specific candidate.
        weight=4 — most common staff action.
        Staff clicks a candidate → files load instantly.
        """
        if not self.cand_ids:
            return
        uid = random.choice(self.cand_ids)
        self.client.get(
            f"/api/admin/candidates/{uid}/files/",
            headers=self.headers,
            name="GET /api/admin/candidates/{uid}/files/",
        )

    @task(1)
    @tag("staff", "read")
    def list_users(self):
        """GET candidate user list — Manage Users page."""
        self.client.get(
            "/api/users/?role=Candidate",
            headers=self.headers,
            name="GET /api/users/?role=Candidate",
        )


# ══════════════════════════════════════════════════════════════
# USER CLASS 3: AdminUser
#
# weight=1 → 10% of virtual users are Admin
# wait_time between(2,8) → admin checks quickly
#
# NOTE: There is only 1 LTADMIN01 account.
# If you spawn many AdminUser VUs they all share the same token.
# In Locust UI, keep total users at 356 (350+5+1) with weight
# 7:2:1 — this naturally spawns ~1 AdminUser.
# ══════════════════════════════════════════════════════════════
class AdminUser(HttpUser):
    weight    = 1
    wait_time = between(2, 8)

    def on_start(self):
        token, _ = do_login(self.client, ADMIN_CREDS)
        if not token:
            raise StopUser()
        self.token   = token
        self.headers = {"Authorization": f"Bearer {token}"}

    def on_stop(self):
        self.client.post(
            "/api/auth/logout/",
            json={"refresh": ""},
            headers=self.headers,
            name="POST /api/auth/logout/ (admin)",
        )

    @task(3)
    @tag("admin", "read")
    def list_all_users(self):
        self.client.get("/api/users/", headers=self.headers, name="GET /api/users/ (all)")

    @task(2)
    @tag("admin", "read")
    def bulk_history(self):
        self.client.get(
            "/api/users/bulk/history/",
            headers=self.headers,
            name="GET /api/users/bulk/history/",
        )

    @task(2)
    @tag("admin", "read")
    def list_labs(self):
        self.client.get("/api/admin/labs/", headers=self.headers, name="GET /api/admin/labs/ (admin)")

    @task(1)
    @tag("admin", "read")
    def check_session(self):
        self.client.get("/api/auth/me/", headers=self.headers, name="GET /api/auth/me/ (admin)")


# ══════════════════════════════════════════════════════════════
# Console summary when test ends
# ══════════════════════════════════════════════════════════════
@events.quitting.add_listener
def on_quit(environment, **kwargs):
    stats = environment.stats
    print("\n" + "=" * 78)
    print("  LOAD TEST COMPLETE — SUMMARY")
    print("=" * 78)
    fmt = "  {:<7} {:<46} {:>6} {:>6} {:>8} {:>8}"
    print(fmt.format("Method", "Endpoint", "Reqs", "Fails", "Avg", "p95"))
    print("-" * 78)
    for name, e in sorted(stats.entries.items(), key=lambda x: -x[1].num_requests):
        if e.num_requests == 0:
            continue
        print(fmt.format(
            e.method,
            name[:45],
            e.num_requests,
            e.num_failures,
            f"{e.avg_response_time:.0f}ms",
            f"{e.get_response_time_percentile(0.95):.0f}ms",
        ))
    print("=" * 78)
    total_reqs  = sum(e.num_requests  for e in stats.entries.values())
    total_fails = sum(e.num_failures  for e in stats.entries.values())
    fail_pct    = (total_fails / total_reqs * 100) if total_reqs else 0
    print(f"\n  Total requests  : {total_reqs}")
    print(f"  Total failures  : {total_fails}  ({fail_pct:.2f}%)")
    result = "PASS ✓" if fail_pct < 2 else "FAIL ✗  (failure rate > 2%)"
    print(f"  Result          : {result}")
    print()