# ================================================================
#  IIAP UploadFiles - Load Test Data Seeder
#
#  HOW TO RUN (from backend folder):
#    cd S:/UploadFiles/backend
#    python loadtest/seed_test_data.py
#
#  Reads from loadtest/test_users.xlsx
#  Columns: SI No | UserID | Name | Role | Password | Mobile | Lab | Action
#
#  Maps roles:
#    Candidate -> Candidate
#    Staff     -> Staff
#    Admin     -> SuperAdmin  (creates as SuperAdmin in DB)
#
#  Safe to re-run -- skips users that already exist.
# ================================================================

import os
import sys

THIS_FILE = os.path.abspath(__file__)
LOADTEST  = os.path.dirname(THIS_FILE)
BACKEND   = os.path.dirname(LOADTEST)

sys.path.insert(0, BACKEND)

env_path = os.path.join(BACKEND, '.env')
if os.path.exists(env_path):
    with open(env_path, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))

os.environ['DJANGO_SETTINGS_MODULE'] = 'config.settings'

import django
django.setup()

from users.models import User  # noqa

RED    = '\033[91m'
GREEN  = '\033[92m'
YELLOW = '\033[93m'
BLUE   = '\033[94m'
BOLD   = '\033[1m'
RESET  = '\033[0m'

def log(sym, color, msg):
    print(color + sym + '  ' + msg + RESET)

def section(title):
    print()
    print(BOLD + BLUE + '-' * 58 + RESET)
    print(BOLD + BLUE + '  ' + title + RESET)
    print(BOLD + BLUE + '-' * 58 + RESET)

# ── Role mapping: Excel value -> Django model role ─────────────
ROLE_MAP = {
    'Candidate': 'Candidate',
    'Staff':     'Staff',
    'Admin':     'SuperAdmin',   # Excel uses 'Admin', DB uses 'SuperAdmin'
}

# ── Load users from xlsx ──────────────────────────────────────
XLSX_PATH = os.path.join(LOADTEST, 'test_users.xlsx')
users_to_create = []

section('Loading test_users.xlsx')

try:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['All Users']
    # Columns: SI No(0) | UserID(1) | Name(2) | Role(3) | Password(4) | Mobile(5) | Lab(6) | Action(7)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1] or row[1] == 'TOTAL':
            continue
        user_id  = str(row[1]).strip()
        name     = str(row[2]).strip() if row[2] else ''
        excel_role = str(row[3]).strip() if row[3] else ''
        password = str(row[4]).strip() if row[4] else 'Test@1234'
        mobile   = str(int(row[5])) if isinstance(row[5], float) else str(row[5] or '')
        lab      = str(row[6]).strip() if row[6] else ''
        db_role  = ROLE_MAP.get(excel_role)
        if not db_role:
            continue   # skip unknown roles
        users_to_create.append({
            'user_id':  user_id,
            'name':     name,
            'role':     db_role,
            'password': password,
            'mobile':   mobile,
            'lab':      lab if lab not in ('', 'All') else None,
        })
    wb.close()
    log('✓', GREEN, 'Loaded ' + str(len(users_to_create)) + ' users from test_users.xlsx')
except Exception as exc:
    log('!', RED, 'Could not read xlsx: ' + str(exc))
    log('!', YELLOW, 'Using inline fallback credentials')
    for i in range(1, 351):
        lab = 'Lab-' + chr(65 + ((i - 1) // 50))
        users_to_create.append({'user_id': 'LTCAND{:03d}'.format(i), 'name': 'LT Candidate {}'.format(i),
                                 'role': 'Candidate', 'password': 'Test@1234', 'mobile': '', 'lab': lab})
    for i in range(1, 150):
        lab = 'Lab-' + chr(65 + ((i - 1) // 21))
        users_to_create.append({'user_id': 'LTSTAFF{:03d}'.format(i), 'name': 'LT Staff {}'.format(i),
                                 'role': 'Staff', 'password': 'Test@1234', 'mobile': '', 'lab': lab})
    users_to_create.append({'user_id': 'LTADMIN01', 'name': 'LT Admin',
                             'role': 'SuperAdmin', 'password': 'Test@1234', 'mobile': '', 'lab': None})

# ── Create users ──────────────────────────────────────────────
section('Creating users in database')

created = 0
skipped = 0
errors  = 0

for spec in users_to_create:
    uid = spec['user_id']
    if User.objects.filter(user_id=uid).exists():
        skipped += 1
        continue
    try:
        User.objects.create_user(
            user_id  = uid,
            password = spec['password'],
            name     = spec['name'],
            role     = spec['role'],
            mobile   = spec['mobile'] or None,
            lab      = spec['lab'],
            created_by = 'LOADTEST_SEEDER',
        )
        created += 1
        if created % 50 == 0:
            log('✓', GREEN, 'Created ' + str(created) + ' users so far...')
    except Exception as exc:
        log('x', RED, uid + ' -- ' + str(exc))
        errors += 1

# ── Summary ───────────────────────────────────────────────────
section('SEED COMPLETE')

total    = len(users_to_create)
cands    = sum(1 for u in users_to_create if u['role'] == 'Candidate')
staff    = sum(1 for u in users_to_create if u['role'] == 'Staff')
admins   = sum(1 for u in users_to_create if u['role'] == 'SuperAdmin')

print()
print('  Defined in Excel  : ' + str(total))
print('    Candidates      : ' + str(cands) + '  (LTCAND001-LTCAND350)')
print('    Staff           : ' + str(staff) + '   (LTSTAFF001-LTSTAFF149)')
print('    Admin           : ' + str(admins) + '    (LTADMIN01)')
print()
log('✓', GREEN,  'Created in DB     : ' + str(created))
if skipped:
    log('-', YELLOW, 'Skipped (existed) : ' + str(skipped))
if errors:
    log('x', RED,    'Errors            : ' + str(errors))
print()

if errors == 0:
    print(BOLD + GREEN + '  All users ready. Start locust now.' + RESET)
    print()
    print('  Locust UI settings:')
    print('    Number of users : 356  (350 candidates + 5 staff + 1 admin)')
    print('    Spawn rate      : 10   (realistic ramp over 35 seconds)')
    print('    Host            : http://localhost:8000')
else:
    print(BOLD + RED + '  Some users failed. Check errors above.' + RESET)
print()